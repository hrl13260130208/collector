import openpyxl


# 写excel
def write_excel():
    wb = openpyxl.Workbook()  # 创建工作簿

    sheet = wb.create_sheet("sheet1", 0)

    sheet.cell(row=1, column= 1).value = '1'

    sheet.cell(row=2, column= 2).value = '2'

    wb.save("chatPy.xlsx")  # 保存文件


if __name__ == '__main__':
    # 写入Excel
    write_excel()
    print('写入成功')
